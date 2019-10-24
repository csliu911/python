'''
Created on 2019年9月19日

@author: liushucheng
'''
# coding=utf-8

import os
import re
import shutil
import openpyxl

msn = ['adc','cortst','dio','eth','fls','flstst','gpt','icu','lin','mcu','port','pwm','ramtst','wdg','general']
MSN = ['ADC','Cortst','DIO','ETH','FLS','Flstst','GPT','ICU','LIN','MCU','PORT','PWM','RamTst','WDG','General']

# msn = ['general']#used only in debug mode
# MSN = ['General']#used only in debug mode

# msn = []#used only in debug mode
# MSN = []#used only in debug mode

# document number location U3
row_num_document_number = 3
col_num_document_number = 21

# 从目标sheet的第三行第三列开始写入
row_num_base = 3
col_num_base = 3

# 记录结果的文件
dstfile = r"C:\Users\liushucheng\Desktop\check_result\Review_DOC_Number_Ver.4.00.05.B.xlsx"

if os.path.exists(dstfile):#文件已经存在
    wb_dstfile = openpyxl.load_workbook(dstfile)
    #returns the list of the names of worksheets in this workbook
    ws_dstfile = wb_dstfile.sheetnames
    print('$$$$' + str(ws_dstfile))
    #return a worksheet by its name
    ws_written = wb_dstfile[ws_dstfile[0]]#get worksheet by name
    print('$$$$' + str(ws_written) + " is to be written to.")
else:#文件不存在
    print('$$$$' + str(dstfile) + " does not exsit.")

delta_i = 0

# main process start
for i in range(len(msn)):
    # 保存review议事录的路径
    path = "U:\\internal\\X1X\\F1x\\modules\\" + msn[i] + "\\review\\ILCD\\F1K_F1KM_Ver4.05.00_Ver42.05.00_F1KH_Ver42.05.00_ASILB\\"
    print(path)
    list_dir = os.listdir(path)
    for j in range(len(list_dir)):
        # print(list_dir[j])
        try:
            wb = openpyxl.load_workbook(path + list_dir[j])
            # 获取workbook中所有的表格
            sheets = wb.sheetnames
            ws_read = wb[sheets[0]] #文书番号在一个review议事录中是相同的,所以只读取第一页的文书番号
            doc_number = ws_read.cell(row = row_num_document_number, column = col_num_document_number).value
        except:
            doc_number = "Load failed"
        print(list_dir[j] + " document number:" + str(doc_number))
        ws_written.cell(row = row_num_base + delta_i, column = col_num_base).value = msn[i]
        ticket_num = re.findall(r'ARDAABD-\d+',str(list_dir[j]))# 匹配 ticket number
        ws_written.cell(row = row_num_base + delta_i, column = col_num_base + 1).value = str(ticket_num[0]) # 只能匹配到一个 ticket number
        ws_written.cell(row = row_num_base + delta_i, column = col_num_base + 2).value = str(doc_number)
        delta_i += 1
# 存放review议事录的路径
path_fusa = "U:\\internal\\X1X\\F1x\\common_family\\docs\\Review\\FUSA\\F1K_F1KM_Ver4.05.00_Ver42.05.00_F1KH_Ver42.05.00_ASILB\\"
print(path_fusa)
list_dir = os.listdir(path_fusa)
for i in range(len(list_dir)):
    try:
        wb = openpyxl.load_workbook(path_fusa + list_dir[i])
        # 获取workbook中所有的表格
        sheets = wb.sheetnames
        ws_read = wb[sheets[0]] #文书番号在一个review议事录中是相同的,所以只读取第一页的文书番号
        doc_number = ws_read.cell(row = row_num_document_number, column = col_num_document_number).value
    except:
        doc_number = "Load failed"
    print(list_dir[i] + " document number:" + str(doc_number))
    ws_written.cell(row = row_num_base + delta_i, column = col_num_base).value = "FUSA"
    ticket_num = re.findall(r'ARDAABD-\d+',str(list_dir[i]))# 匹配 ticket number
    ws_written.cell(row = row_num_base + delta_i, column = col_num_base + 1).value = str(ticket_num[0]) # 从文件名称里只能匹配到一个 ticket number
    ws_written.cell(row = row_num_base + delta_i, column = col_num_base + 2).value = str(doc_number)
    delta_i += 1
# 存放review议事录的路径
path_cov = "U:\\internal\\X1X\\F1x\\common_family\\docs\\FuSa\\Configuration_overview\\Configuration_Overview_Ver4.05.00.B_Ver42.05.00.B\\F1Kx_MCAL_ConfigOverview_Peer_Review_Minutes_ARDAABD-4596.xlsm"
print(path_cov)
try:
    wb = openpyxl.load_workbook(path_cov)
    # 获取workbook中所有的表格
    sheets = wb.sheetnames
    ws_read = wb[sheets[0]] #文书番号在一个review议事录中是相同的,所以只读取第一页的文书番号
    doc_number = ws_read.cell(row = row_num_document_number, column = col_num_document_number).value
except:
    doc_number = "Load failed"
print(path_cov + " document number:" + str(doc_number))
ws_written.cell(row = row_num_base + delta_i, column = col_num_base).value = "COV"
ticket_num = re.findall(r'ARDAABD-\d+',path_cov)# 匹配 ticket number
ws_written.cell(row = row_num_base + delta_i, column = col_num_base + 1).value = str(ticket_num[0]) # 从文件名称里只能匹配到一个 ticket number
ws_written.cell(row = row_num_base + delta_i, column = col_num_base + 2).value = str(doc_number)

# save result file
wb_dstfile.save(dstfile)
print("---process completed---")


