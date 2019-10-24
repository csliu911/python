# coding=utf-8

'''
@说明  这个脚本用来更新ASIL B 发布的日程表
@依赖  需要从JIRA上手动获取每个ticket的状态，更新到ASILB_Release_Schedule.xlsx的sheet1中
'''

import os
import openpyxl
import time

# 定义sheet1中记录ticket的起始行
Ticket_Record_Staring_Row = 3

# 定义一个字典数据类型，用于保存各成果物所在列,成果物在文件ASILB_Release_Schedule.xlsx中
dict_wps_column = {'Module_Name':2,'Ticket_Name':3,'JIRA_Status':7,'ESDD':8,'TSDD':9,'ESTR':10,'TSTR':11,'EUM':12,'TUM':13,'ESTS':14,\
                    'TSTP':15,'ECODE':16,'TCODE':17,'PDF':18,'UTP':19,'UTR':20,'QAC':21,'Reqtify':23,'FMEA':24,'DFA':25,'TEQ':26,'AMDC':27,\
                    'Sample Application':28,'BSWMDT':29}

# 定义sheet1中记录的ticket的总数
Total_Tickets_Number = 124

# 定义ASILB release schedule 表格中成果物列表顺序
list_asilb_release_wps = ['ESDD','TSDD','ESTR','TSTR','EUM','TUM','ESTS','TSTP','ECODE','TCODE','PDF','UTP','UTR','QAC',\
                            'Reqtify','FMEA','DFA','TEQ','AMDC','Sample Application','BSWMDT']

msn = ['adc','can','canV2','cortst','dio','eth','fls','flstst','fr','gpt','icu','lin','mcu','port','pwm','ramtst','spi','wdg','general']
MSN = ['ADC','CAN','CAN','CORTST','DIO','ETH','FLS','FLSTST','FR','GPT','ICU','LIN','MCU','PORT','PWM','RAMTST','SPI','WDG','General']

jira_ticket_stat = ['CLOSED','SOLVED','ACCEPTANCE','TESTING','IMPLEMENTATION','DESIGN','WAITING FIX VERSION','CCB REVIEW','ANALYSIS','OPEN']

schedule_file = r'U:\internal\X1X\common_platform\docs\Impact_analysis\F1Kx_Ver4.05.00_Ver42.05.00_ASILB\ASILB_Release_Schedule.xlsx'

# 定义一个空列表，对程序运行过程中遍历到的模块进行记录
list_module_travelled = []

# 加载schdule_file文件到变量wb_schedule, 属性data_only用于读取单元格的值，而不是公式
wb_schedule = openpyxl.load_workbook(schedule_file,data_only = True)
# 获取工作簿wb_schedule中所有工作表的名称
ws_names = wb_schedule.sheetnames
# 测试读取结果
print('$$$$' + str(ws_names))
# 根据工作表的名称获取工作表
ws_sheet1 = wb_schedule[ws_names[0]]
# 测试读取结果
print('$$$$' + str(ws_sheet1))
ws_schedule = wb_schedule[ws_names[2]]
# 测试读取结果
print('$$$$' + str(ws_schedule))

# 从sheet1中遍历ticket记录
for row_index_inc in range(Total_Tickets_Number):
    # 从module列读取模块名称
    module_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['Module_Name']).value#read cell
    # 测试读取结果
    # print(module_column_cell_data)
    # 读取该模块所在行的ticket列的内容，将作为用于存储这个票的成果物的字典的键值
    ticket_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['Ticket_Name']).value#read cell
    jira_status_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['JIRA_Status']).value#read cell
    ESDD_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['ESDD']).value#read cell
    TSDD_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['TSDD']).value#read cell
    ESTR_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['ESTR']).value#read cell
    TSTR_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['TSTR']).value#read cell
    EUM_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['EUM']).value#read cell
    TUM_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['TUM']).value#read cell
    ESTS_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['ESTS']).value#read cell
    TSTP_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['TSTP']).value#read cell
    ECODE_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['ECODE']).value#read cell
    TCODE_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['TCODE']).value#read cell
    PDF_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['PDF']).value#read cell
    UTP_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['UTP']).value#read cell
    UTR_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['UTR']).value#read cell
    QAC_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['QAC']).value#read cell
    Reqtify_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['Reqtify']).value#read cell
    FMEA_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['FMEA']).value#read cell
    DFA_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['DFA']).value#read cell
    TEQ_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['TEQ']).value#read cell
    AMDC_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['AMDC']).value#read cell
    SAMP_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['Sample Application']).value#read cell
    BSWMDT_column_cell_data = ws_sheet1.cell(row = Ticket_Record_Staring_Row + row_index_inc, column = dict_wps_column['BSWMDT']).value#read cell
    # 创建一个临时列表用于保存成果物的状态信息
    list_wps_status_temp = [jira_status_column_cell_data,ESDD_column_cell_data,TSDD_column_cell_data,ESTR_column_cell_data,\
                            TSTR_column_cell_data,EUM_column_cell_data,TUM_column_cell_data,ESTS_column_cell_data,\
                            TSTP_column_cell_data,ECODE_column_cell_data,TCODE_column_cell_data,PDF_column_cell_data,\
                            UTP_column_cell_data,UTR_column_cell_data,QAC_column_cell_data,Reqtify_column_cell_data,\
                            FMEA_column_cell_data,DFA_column_cell_data,TEQ_column_cell_data,AMDC_column_cell_data,\
                            SAMP_column_cell_data,BSWMDT_column_cell_data]
    # 如果遍历到的模块已经存在于list_module_travelled列表中
    if module_column_cell_data in list_module_travelled:
        if module_column_cell_data == 'ADC':
            # 把模块相关成果物的状态追加到第一次遍历到该模块时创建的字典中
            dict_wps_adc[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_adc)
        elif module_column_cell_data == 'CORTST' or module_column_cell_data == 'Cortst':
            # 把模块相关成果物的状态追加到第一次遍历到该模块时创建的字典中
            dict_wps_cortst[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_cortst)
        elif module_column_cell_data == 'DIO':
            # 把模块相关成果物的状态追加到第一次遍历到该模块时创建的字典中
            dict_wps_dio[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_dio)
        elif module_column_cell_data == 'ETH':
            # 把模块相关成果物的状态追加到第一次遍历到该模块时创建的字典中
            dict_wps_eth[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_eth)
        elif module_column_cell_data == 'FLS':
            # 把模块相关成果物的状态追加到第一次遍历到该模块时创建的字典中
            dict_wps_fls[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_fls)
        elif module_column_cell_data == 'FLSTST' or module_column_cell_data == 'Flstst':
            # 把模块相关成果物的状态追加到第一次遍历到该模块时创建的字典中
            dict_wps_flstst[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_flstst)
        elif module_column_cell_data == 'GPT':
            # 把模块相关成果物的状态追加到第一次遍历到该模块时创建的字典中
            dict_wps_gpt[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_gpt)
        elif module_column_cell_data == 'ICU':
            # 把模块相关成果物的状态追加到第一次遍历到该模块时创建的字典中
            dict_wps_icu[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_icu)
        elif module_column_cell_data == 'LIN':
            # 把模块相关成果物的状态追加到第一次遍历到该模块时创建的字典中
            dict_wps_lin[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_lin)
        elif module_column_cell_data == 'MCU':
            # 把模块相关成果物的状态追加到第一次遍历到该模块时创建的字典中
            dict_wps_mcu[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_mcu)
        elif module_column_cell_data == 'PORT':
            # 把模块相关成果物的状态追加到第一次遍历到该模块时创建的字典中
            dict_wps_port[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_port)
        elif module_column_cell_data == 'PWM':
            # 把模块相关成果物的状态追加到第一次遍历到该模块时创建的字典中
            dict_wps_pwm[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_pwm)
        elif module_column_cell_data == 'RAMTST' or module_column_cell_data == 'RamTst':
            # 把模块相关成果物的状态追加到第一次遍历到该模块时创建的字典中
            dict_wps_ramtst[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_ramtst)
        elif module_column_cell_data == 'WDG':
            # 把模块相关成果物的状态追加到第一次遍历到该模块时创建的字典中
            dict_wps_wdg[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_wdg)
        elif module_column_cell_data == 'General':
            # 把模块相关成果物的状态追加到第一次遍历到该模块时创建的字典中
            dict_wps_general[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_general)
    # 如果遍历到的模块是第一次出现
    else:
        # 测试读取结果,只在第一次遍历到的时候打印输出
        print('$$$$$$$$',module_column_cell_data)
        # 把第一次出现的模块名称记录到遍历列表中
        list_module_travelled.append(module_column_cell_data)
        # print(list_module_travelled)
        if module_column_cell_data == 'ADC':
            # 第一次遍历到该模块时创建一个空字典
            dict_wps_adc = {}
            # 把该模块对应的成果物状态保存到刚刚创建的空字典中，键值对中的键是JIRA ticket名称，值是成果物的状态符号
            dict_wps_adc[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_adc)
        elif module_column_cell_data == 'CORTST' or module_column_cell_data == 'Cortst':
            # 第一次遍历到该模块时创建一个空字典
            dict_wps_cortst = {}
            # 把该模块对应的成果物状态保存到刚刚创建的空字典中，键值对中的键是JIRA ticket名称，值是成果物的状态符号
            dict_wps_cortst[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_cortst)
        elif module_column_cell_data == 'DIO':
            # 第一次遍历到该模块时创建一个空字典
            dict_wps_dio = {}
            # 把该模块对应的成果物状态保存到刚刚创建的空字典中，键值对中的键是JIRA ticket名称，值是成果物的状态符号
            dict_wps_dio[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_dio)
        elif module_column_cell_data == 'ETH':
            # 第一次遍历到该模块时创建一个空字典
            dict_wps_eth = {}
            # 把该模块对应的成果物状态保存到刚刚创建的空字典中，键值对中的键是JIRA ticket名称，值是成果物的状态符号
            dict_wps_eth[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_eth)
        elif module_column_cell_data == 'FLS':
            # 第一次遍历到该模块时创建一个空字典
            dict_wps_fls = {}
            # 把该模块对应的成果物状态保存到刚刚创建的空字典中，键值对中的键是JIRA ticket名称，值是成果物的状态符号
            dict_wps_fls[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_fls)
        elif module_column_cell_data == 'FLSTST' or module_column_cell_data == 'Flstst':
            # 第一次遍历到该模块时创建一个空字典
            dict_wps_flstst = {}
            # 把该模块对应的成果物状态保存到刚刚创建的空字典中，键值对中的键是JIRA ticket名称，值是成果物的状态符号
            dict_wps_flstst[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_flstst)
        elif module_column_cell_data == 'GPT':
            # 第一次遍历到该模块时创建一个空字典
            dict_wps_gpt = {}
            # 把该模块对应的成果物状态保存到刚刚创建的空字典中，键值对中的键是JIRA ticket名称，值是成果物的状态符号
            dict_wps_gpt[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_gpt)
        elif module_column_cell_data == 'ICU':
            # 第一次遍历到该模块时创建一个空字典
            dict_wps_icu = {}
            # 把该模块对应的成果物状态保存到刚刚创建的空字典中，键值对中的键是JIRA ticket名称，值是成果物的状态符号
            dict_wps_icu[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_icu)
        elif module_column_cell_data == 'LIN':
            # 第一次遍历到该模块时创建一个空字典
            dict_wps_lin = {}
            # 把该模块对应的成果物状态保存到刚刚创建的空字典中，键值对中的键是JIRA ticket名称，值是成果物的状态符号
            dict_wps_lin[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_lin)
        elif module_column_cell_data == 'MCU':
            # 第一次遍历到该模块时创建一个空字典
            dict_wps_mcu = {}
            # 把该模块对应的成果物状态保存到刚刚创建的空字典中，键值对中的键是JIRA ticket名称，值是成果物的状态符号
            dict_wps_mcu[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_mcu)
        elif module_column_cell_data == 'PORT':
            # 第一次遍历到该模块时创建一个空字典
            dict_wps_port = {}
            # 把该模块对应的成果物状态保存到刚刚创建的空字典中，键值对中的键是JIRA ticket名称，值是成果物的状态符号
            dict_wps_port[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_port)
        elif module_column_cell_data == 'PWM':
            # 第一次遍历到该模块时创建一个空字典
            dict_wps_pwm = {}
            # 把该模块对应的成果物状态保存到刚刚创建的空字典中，键值对中的键是JIRA ticket名称，值是成果物的状态符号
            dict_wps_pwm[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_pwm)
        elif module_column_cell_data == 'RAMTST' or module_column_cell_data == 'RamTst':
            # 第一次遍历到该模块时创建一个空字典
            dict_wps_ramtst = {}
            # 把该模块对应的成果物状态保存到刚刚创建的空字典中，键值对中的键是JIRA ticket名称，值是成果物的状态符号
            dict_wps_ramtst[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_ramtst)
        elif module_column_cell_data == 'WDG':
            # 第一次遍历到该模块时创建一个空字典
            dict_wps_wdg = {}
            # 把该模块对应的成果物状态保存到刚刚创建的空字典中，键值对中的键是JIRA ticket名称，值是成果物的状态符号
            dict_wps_wdg[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_wdg)
        elif module_column_cell_data == 'General':
            # 第一次遍历到该模块时创建一个空字典
            dict_wps_general = {}
            # 把该模块对应的成果物状态保存到刚刚创建的空字典中，键值对中的键是JIRA ticket名称，值是成果物的状态符号
            dict_wps_general[(ticket_column_cell_data)] = list_wps_status_temp
            # print(dict_wps_general)
# *****************************************************************************
'''schedule_file Schedule sheet contents processing'''
# *****************************************************************************
# 定义Schedule sheet中记录模块名称的起始行
module_in_shedule_start_row = 4
print(time.asctime(time.localtime()))



# end of file
