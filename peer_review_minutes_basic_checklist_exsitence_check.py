'''
Created on 2019年9月19日

@author: liushucheng

@description: 这个脚本目的是检查review 议事录中是否有basic checklist sheet.
'''
# coding=utf-8

import os
import re
import shutil
import openpyxl

msn = ['adc','cortst','dio','eth','fls','flstst','gpt','icu','lin','mcu','port','pwm','ramtst','wdg','general']
MSN = ['ADC','Cortst','DIO','ETH','FLS','Flstst','GPT','ICU','LIN','MCU','PORT','PWM','RamTst','WDG','General']

# msn = ['general']#used only in debug mode
# MSN = ['General']#used only in debug mode

# msn = []#used only in debug mode
# MSN = []#used only in debug mode

# 从目标sheet的第三行第三列开始写入
row_num_base = 3
col_num_base = 3

# BasicChecklist模板的存放路径
templateFile = r"U:\internal\X1X\common_platform\docs\Templates\Review\F1Kx_MCAL_Peer_Review_Minutes_Template.xlsm"
# 记录审查结果的文件
dstfile = r"C:\Users\liushucheng\Desktop\check_result\Review_DOC_basic_checklist_exsitence_check.xlsx"


if os.path.exists(dstfile):#记录审查结果的文件已经存在
    wb_dstfile = openpyxl.load_workbook(dstfile)
    #returns the list of the names of worksheets in this workbook
    ws_dstfile = wb_dstfile.sheetnames
    print('$$$$' + str(ws_dstfile))
    #return a worksheet by its name
    ws_written = wb_dstfile[ws_dstfile[0]]#get worksheet by name
    print('$$$$' + str(ws_written) + " is to be written to.")
else:#记录审查结果的文件不存在
    print('$$$$' + str(dstfile) + " does not exsit.")

delta_i = 0

# main process start
for i in range(len(msn)):
    # 保存review议事录的路径
    path = "U:\\internal\\X1X\\F1x\\modules\\" + msn[i] + "\\review\\ILCD\\F1K_F1KM_Ver4.05.00_Ver42.05.00_F1KH_Ver42.05.00_ASILB\\"
    print(path)
    # 取得目录下所有文件名称
    list_dir = os.listdir(path)
    for j in range(len(list_dir)):
        # print(list_dir[j])
        try:
            wb = openpyxl.load_workbook(path + list_dir[j])
            # 获取workbook中所有的表格
            sheets = wb.sheetnames
            if "Checklist" in sheets:
                status = "available"
            else:
                status = "NA"
        except:
            status = "Load failed. Visual confirmation is necessary." # 文件打开失败的提示，需要手动确认
        print(list_dir[j] + " basic checklist exsitence check result:" + str(status))
        ws_written.cell(row = row_num_base + delta_i, column = col_num_base).value = msn[i]
        ticket_num = re.findall(r'ARDAABD-\d+',str(list_dir[j]))# 匹配 ticket number
        ws_written.cell(row = row_num_base + delta_i, column = col_num_base + 1).value = str(ticket_num[0]) # 从文件名成中只能匹配到一个 ticket number
        ws_written.cell(row = row_num_base + delta_i, column = col_num_base + 2).value = str(status)
        delta_i += 1
# 存放review议事录的路径
path_fusa = "U:\\internal\\X1X\\F1x\\common_family\\docs\\Review\\FUSA\\F1K_F1KM_Ver4.05.00_Ver42.05.00_F1KH_Ver42.05.00_ASILB\\"
print(path_fusa)
list_dir = os.listdir(path_fusa)
for i in range(len(list_dir)):
    try:
        wb = openpyxl.load_workbook(path_fusa + list_dir[i])
        # 获取workbook中所有的表格
        sheets = wb.sheetnames
        if "Checklist" in sheets:
            status = "available" # 表示basic checklist 存在
        else:
            status = "NA" # 表示basic checklist 不存在
    except:
        status = "Load failed. Visual confirmation is necessary." # 文件打开失败的提示，需要手动确认
    print(list_dir[i] + " basic checklist exsitence check result:" + str(status))
    ws_written.cell(row = row_num_base + delta_i, column = col_num_base).value = "FUSA"
    ticket_num = re.findall(r'ARDAABD-\d+',str(list_dir[i]))# 匹配 ticket number
    ws_written.cell(row = row_num_base + delta_i, column = col_num_base + 1).value = str(ticket_num[0]) # 从文件名称里只能匹配到一个 ticket number
    ws_written.cell(row = row_num_base + delta_i, column = col_num_base + 2).value = str(status)
    delta_i += 1
# 存放review议事录的路径
path_cov = "U:\\internal\\X1X\\F1x\\common_family\\docs\\FuSa\\Configuration_overview\\Configuration_Overview_Ver4.05.00.B_Ver42.05.00.B\\"
print(path_cov)
list_dir = os.listdir(path_cov)
# print(list_dir)
for file_index in range(len(list_dir)):
    if "Minutes" in list_dir[file_index]: # 目录下可能存在其他文件，只对review议事录进行basic checklist检查，review议事录的文件名称应该包含关键字"Minutes"
        # print(list_dir[file_index])
        try:
            wb = openpyxl.load_workbook(path_cov + list_dir[file_index])
            # 获取workbook中所有的表格
            sheets = wb.sheetnames
            if "Checklist" in sheets:
                status = "available" # 表示basic checklist 存在
            else:
                status = "NA" # 表示basic checklist 不存在
        except:
            status = "Load failed. Visual confirmation is necessary." # 文件打开失败的提示，需要手动确认
        print(list_dir[file_index] + " basic checklist exsitence check result:" + str(status))
        ws_written.cell(row = row_num_base + delta_i, column = col_num_base).value = "COV"
        ticket_num = re.findall(r'ARDAABD-\d+',path_cov + list_dir[file_index])# 匹配 ticket number
        ws_written.cell(row = row_num_base + delta_i, column = col_num_base + 1).value = str(ticket_num[0]) # 从文件名称里只能匹配到一个 ticket number
        ws_written.cell(row = row_num_base + delta_i, column = col_num_base + 2).value = str(status)
        delta_i += 1
    else:
        print(list_dir[file_index] + " is not a review minutes type file.")
# 保存记录审查结果的文件
wb_dstfile.save(dstfile)
print("---process completed---")

# end of file