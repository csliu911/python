'''
Created on 2019年10月31日

@author: liushucheng

@description: 这个脚本目的是检查review 议事录中的公司名称等信息填写是否正确
            : 只检查在变更管理表中有记录的票的review议事录
'''
# coding=utf-8

import os
import re
import shutil
import openpyxl

# 查询的结果从目标sheet的第三行第三列开始写入
row_num_base = 3
col_num_base = 3

# msn = ['adc','cortst','dio','eth','fls','flstst','gpt','icu','lin','mcu','port','pwm','ramtst','wdg','general']
# MSN = ['ADC','Cortst','DIO','ETH','FLS','Flstst','GPT','ICU','LIN','MCU','PORT','PWM','RamTst','WDG','General']

msn = ['flstst']#used only in debug mode
MSN = ['FLSTST']#used only in debug mode

# 保存查询结果的文件
inspection_result_file = "C:\\Users\\liushucheng\\Desktop\\check_result\\peer_review_minutes_inspection_result.xlsx"


if os.path.exists(inspection_result_file):#记录审查结果的文件已经存在
    wb_dstfile = openpyxl.load_workbook(inspection_result_file)
    #returns a list of the names of worksheets in this workbook
    ws_dstfile = wb_dstfile.sheetnames
    # print('$$$$' + str(ws_dstfile))
    #return a worksheet by its name
    ws_written = wb_dstfile[ws_dstfile[0]]#get worksheet by name
    print('>>>> ' + str(ws_written) + ' in ' + inspection_result_file + ' is to be written to.')
else:#记录审查结果的文件不存在
    print('>>>> ' + str(dstfile) + " does not exsit.")


delta_i = 0

# review 议事录部门名称单元格坐标
row_department_name = 9
col_department_name = 21
# review 议事录公司名称单元格坐标
row_company_name = 7
col_company_name = 21

# review议事录的名称里都包含"2019"，作为查询关键字使用
key_words_1 = "2019"
key_words_2 = "Peer"

# 部门名称里的关键字包含"Software"
key_words_3 = "Software"

PREFIX = "ARDAABD-"

# main process start
for i in range(len(msn)):
    # 变更管理表文件路径
    change_management_file = "U:\\internal\\X1X\\common_platform\\docs\\Impact_analysis\\F1Kx_Ver4.05.00_Ver42.05.00_ASILB\\F1Kx_V4.05.00.B_" + MSN[i] + "_Change_Management.xlsx"
    # 获取工作簿
    wb_cm_file = openpyxl.load_workbook(change_management_file)
    # 获取工作簿中所有的工作表名称，保存的形式为一个列表
    ws_names_of_cm_file = wb_cm_file.sheetnames
    # print(ws_names_of_cm_file)

    # 创建一个空列表用来保存变更管理表中的有效工作表的名称(即不包含cover，revision history等辅助工作表)
    list_ws_names_of_cm_file = [] 
    for ws_index in range(len(ws_names_of_cm_file)):
        # 取得工作表名称前缀是"ARDAABD-"的工作表
        if PREFIX in ws_names_of_cm_file[ws_index]:
            # 保存包含关键字的工作表名称到之前创建的列表中
            list_ws_names_of_cm_file.append(ws_names_of_cm_file[ws_index])
        else:
            pass
    print('>>>> tickets list in ' + str(MSN[i]) + ' change management report: ' + str(list_ws_names_of_cm_file))

    # review议事录路径
    peer_review_minutes_path = "U:\\internal\\X1X\\F1x\\modules\\" + msn[i] + "\\review\\ILCD\\F1K_F1KM_Ver4.05.00_Ver42.05.00_F1KH_Ver42.05.00_ASILB\\"
    # 取得目录下所有文件名称
    list_dir = os.listdir(peer_review_minutes_path)

    # 遍历review议事录路径下的所有文件
    for j in range(len(list_dir)):
        # 遍历变更管理表中的有效ticket名称列表
        for index in range(len(list_ws_names_of_cm_file)):
            # 在review议事录的名称中查询到了变更管理表中记录的要变更的ticket名称
            if list_ws_names_of_cm_file[index] in list_dir[j]:
                try:
                    wb_peer_reivew_minutes = openpyxl.load_workbook(peer_review_minutes_path + list_dir[j],read_only = False,data_only = False,keep_vba = True)
                    # 获取review议事录工作簿中所有的工作表
                    ws_names_of_peer_review_minutes = wb_peer_reivew_minutes.sheetnames
                    # print(">>>> worksheet name list: " + str(ws_names_of_peer_review_minutes))
                    # 遍历review议事录工作簿中所有的工作表
                    for k in range(len(ws_names_of_peer_review_minutes)):
                        # print(">>>> worksheet name: " + ws_names_of_peer_review_minutes[k])
                        # 如果工作表是有效的review工作表(非cover、revision history等辅助工作表)
                        if key_words_1 in ws_names_of_peer_review_minutes[k] or key_words_2 in ws_names_of_peer_review_minutes[k]:
                            # print(ws_names_of_peer_review_minutes[k])
                            try:
                                ws_peer_review_minutes = wb_peer_reivew_minutes[ws_names_of_peer_review_minutes[k]]
                                # read cell
                                department_name = ws_peer_review_minutes.cell(row = row_department_name, column = col_department_name).value
                                # print(MSN[i] + ',' + list_dir[j] + ',' + ws_names_of_peer_review_minutes[k] + ',department name,' + department_name)
                                if key_words_3 in department_name:
                                    print(MSN[i] + ',' + list_dir[j] + ',' + ws_names_of_peer_review_minutes[k] + ',department name,' + department_name)
                                else:
                                    # write cell, modify department name if the department name is incorrect(openpyxl 3.0 and later works)
                                    ws_peer_review_minutes.cell(row = row_department_name, column = col_department_name).value = 'Automotive MCU Software Department'
                                    # read cell
                                    department_name = ws_peer_review_minutes.cell(row = row_department_name, column = col_department_name).value
                                    print(MSN[i] + ',' + list_dir[j] + ',' + ws_names_of_peer_review_minutes[k] + ',department name,' + department_name + ', department name modified')
                            except:
                                print(">>>> worksheet name: " + ws_names_of_peer_review_minutes[k] + " department name read failed.")
                        else:
                            pass
                        wb_peer_reivew_minutes.save(peer_review_minutes_path + list_dir[j])
                        # wb_peer_reivew_minutes.save(list_dir[j])
                        wb_peer_reivew_minutes.close()
                except:
                    status = ">>>> Load failed. Visual confirmation is necessary." # 文件打开失败的提示，需要手动确认
                    print(status)
            # 在review议事录的名称中没有查询到变更管理表中记录的要变更的ticket名称
            else:
                pass
        # print(list_dir[j] + " basic checklist exsitence check result:" + str(status))
        # ws_written.cell(row = row_num_base + delta_i, column = col_num_base).value = msn[i]
        # ticket_num = re.findall(r'ARDAABD-\d+',str(list_dir[j]))# 匹配 ticket number