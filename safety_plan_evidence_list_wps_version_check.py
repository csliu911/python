'''
Created on 2019年10月21日
@author: liushucheng
@brief: 检查safetyplan证据列表中记录的成果物版本信息是否和文件的真实信息一致
      : 首先通过批处理文件更新所有成果物，之后用脚本获取成果物的版本号和SVN revision信息
      : 再通过下面的脚本文件从safety plan attachment 中读取记录的成果物信息，做比较
'''
# coding=utf-8
import re
import openpyxl

# *********************************************************************************************************************
# 
# *********************************************************************************************************************
# safetyplan list of evidence 文件所在位置
safety_plan_attachment_srcfile = "U:\\internal\\X1X\\F1x\\common_family\\docs\\FuSa\\Safety_Plan_Ver4.05.00.B_Ver42.05.00.B\\RH850_F1KMH_MCAL_Ver4.05.00.B_Ver42.05.00.B_SafetyPlan_Attachment.xlsm"

# 成果物所有者所在列的序号
col_number_owner = 11
# 成果物文件名称所在列的序号
col_number_wps_name = 17
# 成果物文件路径所在列的序号
col_number_wps_location = 18
# 成果物版本信息所在列的序号
col_number_wps_version = 22

# 以只读方式加载工作簿
wb = openpyxl.load_workbook(safety_plan_attachment_srcfile,data_only = True)

# 根据工作表名称取得所有工作表
ws_names = wb.sheetnames

# 根据工作表序号取得工作表
ws_list_of_evidence = wb[ws_names[2]]

# 获取工作表的行数
row_number_total = ws_list_of_evidence.max_row
# print(row_number_total)

# 成果物版本信息和SVN revision信息的保存文件,这个文件可以通过运行get_doc_num_version_svn_revision4safety_plan_attachment.py
# 的时候，利用输出重定向功能直接生成
result_file_loc = "U:\\internal\\X1X\\F1x\\common_family\\docs\\FuSa\\Safety_Plan_Ver4.05.00.B_Ver42.05.00.B\\extracted_wps_version.txt"

for row_index in range(row_number_total):
    # open file
    result_file = open(result_file_loc,'r',encoding='utf-8')
    # print(row_index)
    # 从safety plan 证据列表中读取成果物的
    owner_department_info = ws_list_of_evidence.cell(row = row_index + 1,column = col_number_owner).value
    # print(owner_department_info)
    # 筛选ILCD负责的成果物
    if 'ILCD' in str(owner_department_info):
        # print('row number,' + str(row_index) + ',' + str(owner_department_info))
        # 读取成果物文件名
        wps_name = ws_list_of_evidence.cell(row = row_index + 1,column = col_number_wps_name).value
        # 读取成果物版本号或者SVN revision
        wps_version = ws_list_of_evidence.cell(row = row_index + 1,column = col_number_wps_version).value
        # 读取成果物文件路径
        wps_location = ws_list_of_evidence.cell(row = row_index + 1,column = col_number_wps_location).value
        # 成果物文件路径中可能包含回车符，形式是‘_x000D_’，需要替换掉
        if "_x000D_" in str(wps_location):
            # print(wps_location + ",origin")
            wps_location = wps_location.replace('_x000D_','')
            # print(wps_location + ",replaced")
        else:
            pass
        # 路径末尾可能包含一个或者多个换行符号，所以下面的语句不是完全适用
        # if '\n' in wps_location:
        #     wps_location = wps_location.replace('\n','')
        # else:
        #     pass
        # 用正则表达式的贪婪匹配模式匹配成果物文件名中的空格符号，替换为空字符
        wps_name = re.sub(r'^\s+','',wps_name)
        # 用正则表达式的贪婪匹配模式匹配成果物文件路径中的换行符号，替换为空字符
        wps_location = re.sub(r'\n+','',wps_location)
        # 用正则表达式在成果物文件路径末尾匹配分隔符"/",如果没有匹配到
        if re.search(r'\/$',wps_location) == None:
            # 在成果物文件路径末尾追加一个分隔符"/"
            wps_location = wps_location + '/'
        # 构造成果物全路径
        wps_full_path = wps_location + wps_name
        # print(wps_full_path)

        # 遍历抽取的成果物版本信息文件
        for line,value in enumerate(result_file):
            # 如果safety plan中的成果物名称包含在 value 中
            if wps_full_path in str(value):
                # 判断safety plan中的成果物版本信息是否和从成果物文件中抽取的版本信息一致
                if wps_version in str(value):
                    print(wps_name + ',Version in Safety Plan :' + wps_version + ',========== Version Abstracted:' + value + ', UP TO DATE.')
                else:
                    print(wps_name + ',Version in Safety Plan :' + wps_version + ',<><><><><> Version Abstracted:' + value + ', OLD VERSION INFO.')
            else:
                pass
    # ILCD负责范围之外的文件不做处理
    else:
        pass
    # 查询完一条记录就需要关闭文件，下次查询再重新打开，否则用于记录遍历位置的变量不会重置
    result_file.close()
    # print(line)


