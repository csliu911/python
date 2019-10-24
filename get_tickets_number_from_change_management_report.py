# coding=utf-8
'''
the program is used to acquire sheet name(ticket number) in change management document,
and acquire the impact on functional safety status and restore at the
Impact_on_FUSA document.
'''

import os
import openpyxl

PREFIX = "ARDAABD-"
row_num = 21
col_num = 4

col_num_ws_fusa = 2
row_base_num_ws_fusa = 3

#author location in change managemant document 
col_num_author = 5
row_num_author = 9
#confirm status by Safety Manager cell location
col_num_SM_confirm = 5
row_num_SM_confirm = 29
#work products to be changed cell location
col_num_wps_changed = 4
row_num_wps_changed = 15


msn = ['adc','can','canV2','cortst','dio','eth','fls','flstst','fr','gpt','icu','lin','mcu','port','pwm','ramtst','spi','wdg','general']
MSN = ['ADC','CAN','CAN','Cortst','DIO','ETH','FLS','Flstst','FR','GPT','ICU','LIN','MCU','PORT','PWM','RamTst','SPI','WDG','General']

# msn = ['adc']#used only in debug mode
# MSN = ['ADC']#used only in debug mode

resultfile = r"U:\internal\X1X\common_platform\docs\Impact_analysis\F1Kx_Ver4.05.00_Ver42.05.00_ASILB\Impact on FUSA.xlsx"

if os.path.exists(resultfile):#Impact on FUSA.xlsx 文件已经存在
    wb_impact_on_fusa = openpyxl.load_workbook(resultfile)
    #returns the list of the names of worksheets in this workbook
    ws_impact_on_fusa = wb_impact_on_fusa.sheetnames
    print('$$$$' + str(ws_impact_on_fusa))
    #return a worksheet by its name
    ws_name_impact_on_fusa = wb_impact_on_fusa[ws_impact_on_fusa[0]]#get worksheet by name
    print('$$$$' + str(ws_name_impact_on_fusa))
else:#Impact on FUSA.xlsx 文件不存在
    wb = openpyxl.Workbook()#生成一个 Workbook 的实例化对象，wb即代表一个工作簿（一个 Excel 文件）
    wb.save(resultfile)
    wb_impact_on_fusa = openpyxl.load_workbook(resultfile)
    #returns the list of the names of worksheets in this workbook
    ws_impact_on_fusa = wb_impact_on_fusa.sheetnames
    print('$$$$' + str(ws_impact_on_fusa))
    #return a worksheet by its name
    ws_name_impact_on_fusa = wb_impact_on_fusa[ws_impact_on_fusa[0]]#get worksheet by name
    print('$$$$' + str(ws_name_impact_on_fusa))

delta_i = 0

for wb_index in range(len(msn)):    
    dstfile = "U:\\internal\\X1X\\common_platform\\docs\\Impact_analysis\\F1Kx_Ver4.05.00_Ver42.05.00_ASILB\\F1Kx_V4.05.00.B_" + MSN[wb_index] + "_Change_Management.xlsx"
    dstfile_exist = os.path.exists(dstfile)
    if not dstfile_exist:
        print("$$$$destination workbook does not exist, please check file name ",dstfile)
    else:
        print("$$$$destination workbook exist ",dstfile)
        wb = openpyxl.load_workbook("U:\\internal\\X1X\\common_platform\\docs\\Impact_analysis\\F1Kx_Ver4.05.00_Ver42.05.00_ASILB\\F1Kx_V4.05.00.B_" + MSN[wb_index] + "_Change_Management.xlsx")
        print(MSN[wb_index])
        #获取workbook中所有的表格
        #提取名称里包含ARDAABD的项目
        sheets = wb.sheetnames
        for ws_index in range(len(sheets)):
            if PREFIX in sheets[ws_index]:
                print(sheets[ws_index])
                table = wb[sheets[ws_index]]#get worksheet by name
                #print(table)
                #save module name to document impact_on_fusa at column 2
                ws_name_impact_on_fusa.cell(row = row_base_num_ws_fusa + delta_i, column = col_num_ws_fusa).value = MSN[wb_index]#write cell
                #save ticket name to document impact_on_fusa at column 3
                ws_name_impact_on_fusa.cell(row = row_base_num_ws_fusa + delta_i, column = col_num_ws_fusa+1).value = sheets[ws_index]#write cell
                #acquire the status of impact on functional safety which is detailed at cell D21
                cell_data = table.cell(row = row_num, column = col_num).value#read cell
                #acquire author name
                author_name = table.cell(row = row_num_author, column = col_num_author).value#read cell
                #acquire Safety Manager confirmation status
                sm_confirm_status = table.cell(row = row_num_SM_confirm, column = col_num_SM_confirm).value#read cell
                #acquire work products to be changed contents
                wps_changed = table.cell(row = row_num_wps_changed, column = col_num_wps_changed).value#read cell
                #save impact on functional safety status to document impact_on_fusa at column 4
                ws_name_impact_on_fusa.cell(row = row_base_num_ws_fusa + delta_i, column = col_num_ws_fusa+2).value = cell_data#write cell
                # save author name at column 5
                ws_name_impact_on_fusa.cell(row = row_base_num_ws_fusa + delta_i, column = col_num_ws_fusa+3).value = author_name#write cell
                # save Safety Manager confirm at column 6
                ws_name_impact_on_fusa.cell(row = row_base_num_ws_fusa + delta_i, column = col_num_ws_fusa+4).value = sm_confirm_status#write cell
                # save work products to be changed at column 8
                ws_name_impact_on_fusa.cell(row = row_base_num_ws_fusa + delta_i, column = col_num_ws_fusa+5).value = wps_changed#write cell
                delta_i += 1#change cell row index
# 保存记录结果的文件
wb_impact_on_fusa.save(resultfile)
print("---process completed---")
                
# end of file