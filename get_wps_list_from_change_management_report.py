'''
Created on 2019年9月7日

@author: liushucheng
'''
# coding=utf-8
import os
import openpyxl

PREFIX = "ARDAABD-"

msn = ['adc','cortst','dio','eth','fls','flstst','gpt','icu','lin','mcu','port','pwm','ramtst','wdg','general']
MSN = ['ADC','Cortst','DIO','ETH','FLS','Flstst','GPT','ICU','LIN','MCU','PORT','PWM','RamTst','WDG','General']

# msn = ['general']#used only in debug mode
# MSN = ['General']#used only in debug mode

row_num_base = 3
col_num_base = 2

col_num_wps_effort = 4 #column D, effort of each work product is listed here
row_num_wps_effort = 22 #row 22, effort of each work product is listed here

col_num_approved_pl = 4#project leader approved status
row_num_approved_pl = 29#project leader approved status

col_num_approved_sm = 5#safety manager approved status
row_num_approved_sm = 29#safety manager approved status

resultfile = r"U:\internal\X1X\common_platform\docs\Impact_analysis\F1Kx_Ver4.05.00_Ver42.05.00_ASILB\wps_schedule_of_tickets.xlsx"

if os.path.exists(resultfile):#文件已经存在
    wb_wps_of_tickets = openpyxl.load_workbook(resultfile)
    #returns the list of the names of worksheets in this workbook
    ws_wps_of_tickets = wb_wps_of_tickets.sheetnames
    print('$$$$' + str(ws_wps_of_tickets))
    #return a worksheet by its name
    ws_written = wb_wps_of_tickets[ws_wps_of_tickets[0]]#get worksheet by name
    print('$$$$' + str(ws_written))
else:#文件不存在
    wb = openpyxl.Workbook()#生成一个 Workbook 的实例化对象，wb即代表一个工作簿（一个 Excel 文件）
    wb.save(resultfile)
    wb_wps_of_tickets = openpyxl.load_workbook(resultfile)
    #returns the list of the names of worksheets in this workbook
    ws_wps_of_tickets = wb_wps_of_tickets.sheetnames
    print('$$$$' + str(ws_wps_of_tickets))
    #return a worksheet by its name
    ws_written = wb_wps_of_tickets[ws_wps_of_tickets[0]]#get worksheet by name
    print('$$$$' + str(ws_written))

delta_i = 0

for wb_index in range(len(msn)):
    dstfile = "U:\\internal\\X1X\\common_platform\\docs\\Impact_analysis\\F1Kx_Ver4.05.00_Ver42.05.00_ASILB\\F1Kx_V4.05.00.B_" + MSN[wb_index] + "_Change_Management.xlsx"
    dstfile_exist = os.path.exists(dstfile)
    if not dstfile_exist:
        print("$$$$destination workbook does not exist, please check file name ",dstfile)
    else:
        print("$$$$destination workbook exist ",dstfile)
        wb = openpyxl.load_workbook("U:\\internal\\X1X\\common_platform\\docs\\Impact_analysis\\F1Kx_Ver4.05.00_Ver42.05.00_ASILB\\F1Kx_V4.05.00.B_" + MSN[wb_index] + "_Change_Management.xlsx")
        print(MSN[wb_index])
        # 获取workbook中所有的表格
        sheets = wb.sheetnames
        # 提取名称里包含ARDAABD的sheet
        for ws_index in range(len(sheets)):
            # 取得所有前缀是"ARDAABD-"的tickets
            if PREFIX in sheets[ws_index]:
                # print(sheets[ws_index])
                ws_read = wb[sheets[ws_index]]#get worksheet by name
                print(ws_read)
                #save module name at column 2
                ws_written.cell(row = row_num_base + delta_i, column = col_num_base).value = MSN[wb_index]#write cell
                #save ticket name at column 3
                ws_written.cell(row = row_num_base + delta_i, column = col_num_base+1).value = sheets[ws_index]#write cell
                # read cell D22
                wps_effort_cell_data = ws_read.cell(row = row_num_wps_effort, column = col_num_wps_effort).value#read cell D22
                # print(wps_effort_cell_data)
                ws_written.cell(row = row_num_base + delta_i, column = col_num_base+2).value = wps_effort_cell_data#write cell
                #read cell D29
                approved_status_pl = ws_read.cell(row = row_num_approved_pl, column = col_num_approved_pl).value#read cell D29
                ws_written.cell(row = row_num_base + delta_i, column = col_num_base+3).value = approved_status_pl#write cell
                #read cell E29
                approved_status_sm = ws_read.cell(row = row_num_approved_sm, column = col_num_approved_sm).value#read cell E29
                ws_written.cell(row = row_num_base + delta_i, column = col_num_base+4).value = approved_status_sm#write cell
                delta_i += 1#change cell row index
        # 保存
        wb_wps_of_tickets.save(resultfile)
print("---process completed---")

