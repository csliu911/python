'''
Created on 2019年9月5日

@author: liushucheng
'''
# coding=utf-8
import os
import time
import openpyxl

PREFIX = "ARDAABD-"

# case sensitive
key_word_1 = "ECODE"
key_word_2 = "PDF"
key_word_3 = 'Yes'
key_word_4 = 'YES'

# revision number mark
RevPrefix = "Rev. "

rev_history_author = 'Shucheng Liu'
rev_history_comment = 'Justification for no need to implement FMEA activities was appended to impacts on functional safety.'

# This content is to be added to cell D21 of change management sheet for tickets whose source code had been modified.
txt = "Although the source code was modified, the analysis showed that the modification was not FSR related. Therefore, there is no need to implement FMEA activities."

#msn = ['adc','can','canV2','cortst','dio','eth','fls','flstst','fr','gpt','icu','lin','mcu','port','pwm','ramtst','spi','wdg','general']
#MSN = ['ADC','CAN','CAN','Cortst','DIO','ETH','FLS','Flstst','FR','GPT','ICU','LIN','MCU','PORT','PWM','RamTst','SPI','WDG','General']

msn = ['adc','cortst','dio','eth','fls','flstst','gpt','icu','lin','mcu','port','pwm','ramtst','wdg','general']
MSN = ['ADC','Cortst','DIO','ETH','FLS','Flstst','GPT','ICU','LIN','MCU','PORT','PWM','RamTst','WDG','General']

# msn = ['general']#used only in debug mode
# MSN = ['General']#used only in debug mode

col_num_impact_on_fusa = 4 #column D, impact on functional safety
row_num_impact_on_fusa = 21 #row 21, impact on functional safety

col_num_wps2be_changed = 4 #column D, Name and version of work products to be changed
row_num_wps2be_changed = 15 #row 15, Name and version of work products to be changed

# tickets_implemented_fmea = ['4412','2787','2513','3837','4442','3442','3388','4061','4075','3817','4347','3997','2938','2705']
tickets_implemented_fmea = [PREFIX + '4412',PREFIX + '2787',PREFIX + '2513',PREFIX + '3837',PREFIX + '4442',PREFIX + '3442',PREFIX + '3388',PREFIX + '4061',PREFIX + '4075',PREFIX + '3817',PREFIX + '4347',PREFIX + '3997',PREFIX + '2938',PREFIX + '2705']
# tickets_sub_4393 = [PREFIX + '4574',PREFIX + '4587',PREFIX + '4586',PREFIX + '4575',PREFIX + '4583',PREFIX + '4580',PREFIX + '4579',PREFIX + '4578',PREFIX + '4576',PREFIX + '4582',PREFIX + '4393',PREFIX + '4581',PREFIX + '4584',PREFIX + '4585']

col_num_rev_base = 2 #column B, revison history starting cell
row_num_rev_base = 6 #row 6, revison history starting cell

# def revision_history_update():

delta_i = 0

for wb_index in range(len(msn)):
    dstfile = "U:\\internal\\X1X\\common_platform\\docs\\Impact_analysis\\F1Kx_Ver4.05.00_Ver42.05.00_ASILB\\F1Kx_V4.05.00.B_" + MSN[wb_index] + "_Change_Management.xlsx"
    dstfile_exist = os.path.exists(dstfile)
    if not dstfile_exist:
        print("$$$$destination workbook does not exist, please check file name ",dstfile)
    else:
        print("$$$$destination workbook exist ",dstfile)
        wb = openpyxl.load_workbook("U:\\internal\\X1X\\common_platform\\docs\\Impact_analysis\\F1Kx_Ver4.05.00_Ver42.05.00_ASILB\\F1Kx_V4.05.00.B_" + MSN[wb_index] + "_Change_Management.xlsx")
        print(MSN[wb_index])
        # 打开新的工作簿后清空之前的tickets记录
        rev_history_sheet = None
        #获取workbook中所有的表格
        sheets = wb.sheetnames
        #提取名称里包含ARDAABD的sheet
        for ws_index in range(len(sheets)):
            # 取得不需要实施FMEA的tickets
            if PREFIX in sheets[ws_index] and sheets[ws_index] not in tickets_implemented_fmea:
                # print(sheets[ws_index])
                table = wb[sheets[ws_index]]#get worksheet by name
                print(table)
                # read work products to be changed data detailed at cell D15
                wps2be_changed_cell_data = table.cell(row = row_num_wps2be_changed, column = col_num_wps2be_changed).value#read cell D15
                # print(wps2be_changed_cell_data)
                # 首先判断变更内容是否涉及ECODE，只在ECODE(包含PDF)有变更的情况下才追加不需要实施FMEA的理由
                if key_word_1 in wps2be_changed_cell_data or key_word_2 in wps2be_changed_cell_data:
                    #acquire the status of impact on functional safety which is detailed at cell D21
                    impact_cell_data = table.cell(row = row_num_impact_on_fusa, column = col_num_impact_on_fusa).value#read cell D21
                    # print(impact_cell_data)
                    # 构造要写入D21单元格的内容
                    impact_cell_data = impact_cell_data + '\n' + txt
                    # 写入构造的内容到D21(MCU, GPT, ETH, LIN, RAMTST, General 的报告中内嵌了文件，不能使用脚本)
                    #table.cell(row = row_num_impact_on_fusa, column = col_num_impact_on_fusa).value = impact_cell_data#write cell D21
                    print('Tips: ECODE modified for ticket ' + sheets[ws_index] + '. Justification for no need to implement FMEA activities was appended.')
                    # 保存追加了justification的tickets number
                    if rev_history_sheet == None:
                        rev_history_sheet = sheets[ws_index]
                    else:
                        rev_history_sheet = rev_history_sheet + '\n' + sheets[ws_index]
                # tickets变更内容不涉及ECODE
                else:
                    print('Tips: No ECODE modified for ticket ' + sheets[ws_index] + '.')
            # 需要实施FMEA的tickets
            elif sheets[ws_index] in tickets_implemented_fmea:
                # print(sheets[ws_index])
                table = wb[sheets[ws_index]]#get worksheet by name
                print(table)
                print('Tips: FMEA activities implemented for ticket ' + sheets[ws_index] + '. Justification Unnecessary.')
        # 更新修订历史记录
        print('Tips: Justified tickets below' + '\n' + rev_history_sheet)
        RevHistory_sheet = wb['RevHistory']#get worksheet by name
        # print(RevHistory_sheet)
        for delta_i in range(99):
            rev_history = RevHistory_sheet.cell(row = row_num_rev_base + delta_i, column = col_num_rev_base).value#read cell B6
            # print(rev_history)
            # looking for the latest record
            if rev_history == None:
                # appending a new revision record
                if delta_i < 10:
                    RevHistory_sheet.cell(row = row_num_rev_base + delta_i, column = col_num_rev_base).value = RevPrefix + '1.0' + str(delta_i)#version less than 10
                else:
                    RevHistory_sheet.cell(row = row_num_rev_base + delta_i, column = col_num_rev_base).value = RevPrefix + '1.' + str(delta_i)#version more than 10
                RevHistory_sheet.cell(row = row_num_rev_base + delta_i, column = col_num_rev_base + 1).value = rev_history_author#author
                RevHistory_sheet.cell(row = row_num_rev_base + delta_i, column = col_num_rev_base + 2).value = rev_history_sheet#sheet name modified
                RevHistory_sheet.cell(row = row_num_rev_base + delta_i, column = col_num_rev_base + 3).value = rev_history_comment#comment
                RevHistory_sheet.cell(row = row_num_rev_base + delta_i, column = col_num_rev_base + 4).value = time.strftime('%Y-%m-%d',time.localtime(time.time()))#date
                print("Tips: Revision number was updated to " + RevHistory_sheet.cell(row = row_num_rev_base + delta_i, column = col_num_rev_base).value)
                break
            else:
                delta_i += 1
        # 保存
        wb.save(dstfile)
print("---process completed---")
# print(time.time())
# print(time.localtime(time.time()))
print(time.strftime('%Y-%m-%d',time.localtime(time.time())))